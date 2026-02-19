"""
MERCADO duBAIRRO — Script de Processamento de Dados
Transforma os dados brutos do ERP em tabelas otimizadas para o Power BI.

Fontes de entrada:
  1. categoria_analisedevendas_[MES][ANO].xlsx
  2. produtopordia_analisedevendas_[MES][ANO].xlsx
  3. curvaA_analisedevendas_[MES][ANO].xlsx
  4. mesamesproduto2025_analisedevendas.xlsx (histórico)

Saída:
  Base_PowerBI.xlsx com 6 abas (tabelas fato + dimensões)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from collections import defaultdict
from datetime import datetime, timedelta
import os
import sys
import glob
import re

# ============================================================
# CONFIGURAÇÕES
# ============================================================
CUSTO_FIXO = 16913.46
META_LIQUIDA = 0.15
DIAS_OPERACAO_MES = 27  # média
LIMIAR_GIRO_ALTO = 0.6  # 60% dos dias
LIMIAR_MARGEM_ALTA = 50  # 50%
LIMIAR_RUPTURA_GIRO = 0.7  # 70% dos dias = giro diário
LIMIAR_RUPTURA_DIAS = 2  # 2 dias sem venda = alerta
LIMIAR_EROSAO = 3  # 3 pontos percentuais

# Meses em português para detecção
MESES_PT = {
    'janeiro': 1, 'fevereiro': 2, 'março': 3, 'abril': 4,
    'maio': 5, 'junho': 6, 'julho': 7, 'agosto': 8,
    'setembro': 9, 'outubro': 10, 'novembro': 11, 'dezembro': 12,
    'jan': 1, 'fev': 2, 'mar': 3, 'abr': 4,
    'mai': 5, 'jun': 6, 'jul': 7, 'ago': 8,
    'set': 9, 'out': 10, 'nov': 11, 'dez': 12
}

# Feriados nacionais 2026 (ajustar conforme necessidade)
FERIADOS_2026 = [
    datetime(2026, 1, 1),   # Ano Novo
    datetime(2026, 2, 16),  # Carnaval
    datetime(2026, 2, 17),  # Carnaval
    datetime(2026, 4, 3),   # Sexta-feira Santa
    datetime(2026, 4, 21),  # Tiradentes
    datetime(2026, 5, 1),   # Dia do Trabalho
    datetime(2026, 6, 4),   # Corpus Christi
    datetime(2026, 9, 7),   # Independência
    datetime(2026, 10, 12), # N.S. Aparecida
    datetime(2026, 11, 2),  # Finados
    datetime(2026, 11, 15), # Proclamação da República
    datetime(2026, 12, 25), # Natal
]
FERIADOS_SET = set(f.strftime('%Y-%m-%d') for f in FERIADOS_2026)

# Styling
HEADER_FILL = PatternFill('solid', fgColor='2D2D2D')
HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=10)
DATA_FONT = Font(name='Arial', size=10)
YELLOW_FILL = PatternFill('solid', fgColor='FFC107')
GREEN_FILL = PatternFill('solid', fgColor='27AE60')
RED_FILL = PatternFill('solid', fgColor='E74C3C')
LIGHT_GRAY_FILL = PatternFill('solid', fgColor='F5F5F5')
THIN_BORDER = Border(
    left=Side(style='thin', color='DDDDDD'),
    right=Side(style='thin', color='DDDDDD'),
    top=Side(style='thin', color='DDDDDD'),
    bottom=Side(style='thin', color='DDDDDD')
)


def parse_br_number(s):
    """Parse Brazilian number format (1.234,56 or 1.234,56%)"""
    if s is None:
        return 0.0
    s = str(s).strip().replace('%', '')
    if not s or s == '-':
        return 0.0
    s = s.replace('.', '').replace(',', '.')
    try:
        return float(s)
    except ValueError:
        return 0.0


def detect_month_year(filename):
    """Detect month and year from filename like 'categoria_analisedevendas_jan2026.xlsx'"""
    base = os.path.basename(filename).lower().replace('.xlsx', '')
    for mes_name, mes_num in sorted(MESES_PT.items(), key=lambda x: -len(x[0])):
        if mes_name in base:
            year_match = re.search(r'(\d{4})', base)
            if year_match:
                return mes_num, int(year_match.group(1))
    return None, None


def style_header(ws, row=1):
    """Apply header styling to first row"""
    for cell in ws[row]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER


def style_data_rows(ws, start_row=2):
    """Apply alternating row colors and borders"""
    for row_idx, row in enumerate(ws.iter_rows(min_row=start_row, max_row=ws.max_row)):
        for cell in row:
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            if row_idx % 2 == 0:
                cell.fill = LIGHT_GRAY_FILL


def auto_width(ws):
    """Auto-adjust column widths"""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 40)


# ============================================================
# 1. PROCESSAR CATEGORIAS
# ============================================================
def processar_categorias(filepath, mes, ano):
    """Processa arquivo de categorias e retorna dados estruturados"""
    wb = openpyxl.load_workbook(filepath)
    ws = wb['Main sheet']
    
    categorias = []
    total = None
    
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=False):
        nome = row[0].value
        if nome is None or nome == '':
            continue
            
        data = {
            'Mes': mes,
            'Ano': ano,
            'Periodo': f"{mes:02d}/{ano}",
            'Categoria': str(nome).strip(),
            'Qtde_Venda': parse_br_number(row[1].value),
            'Qtde_Documentos': parse_br_number(row[2].value),
            'Vlr_Acrescimos': parse_br_number(row[3].value),
            'Vlr_Descontos': parse_br_number(row[4].value),
            'Ticket_Medio': parse_br_number(row[5].value),
            'Vlr_Venda': parse_br_number(row[6].value),
            'Part_Venda': parse_br_number(row[7].value),
            'Markdown_Pct': parse_br_number(row[8].value),
            'Markdown_Ult_Entrada': parse_br_number(row[9].value),
            'Markup_Pct': parse_br_number(row[10].value),
            'Markup_Ult_Entrada': parse_br_number(row[11].value),
            'Vlr_Lucro': parse_br_number(row[12].value),
            'Part_Lucro': parse_br_number(row[13].value),
            'Custo_Medio_Liq': parse_br_number(row[14].value),
            'Custo_Ult_Entrada_Liq': parse_br_number(row[15].value),
        }
        
        if nome == 'Total':
            total = data
        else:
            categorias.append(data)
    
    wb.close()
    return categorias, total


# ============================================================
# 2. PROCESSAR PRODUTO POR DIA
# ============================================================
def processar_produto_dia(filepath, mes, ano):
    """Processa arquivo de produto por dia e retorna vendas diárias"""
    wb = openpyxl.load_workbook(filepath)
    ws = wb['Main sheet']
    
    vendas_diarias = []
    current_date = None
    
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=False):
        col_a = row[0].value
        col_b = row[1].value
        
        if col_a is not None and col_a != 'Total':
            if isinstance(col_a, str) and '/' in col_a:
                parts = col_a.split('/')
                current_date = f"{int(parts[2])}-{int(parts[1]):02d}-{int(parts[0]):02d}"
            else:
                current_date = str(col_a)
        elif col_b is not None and 'Total' not in str(col_b):
            prod_parts = str(col_b).split('||')
            prod_name = prod_parts[0].strip()
            prod_code = prod_parts[1].strip() if len(prod_parts) > 1 else ''
            prod_id = prod_parts[2].strip() if len(prod_parts) > 2 else ''
            
            vlr_venda = parse_br_number(row[7].value)
            vlr_lucro = parse_br_number(row[13].value)
            margem = (vlr_lucro / vlr_venda * 100) if vlr_venda > 0 else 0
            
            vendas_diarias.append({
                'Data': current_date,
                'Mes': mes,
                'Ano': ano,
                'Periodo': f"{mes:02d}/{ano}",
                'Produto': prod_name,
                'Codigo': prod_code,
                'ID_ERP': prod_id,
                'Qtde_Venda': parse_br_number(row[2].value),
                'Qtde_Documentos': parse_br_number(row[3].value),
                'Vlr_Acrescimos': parse_br_number(row[4].value),
                'Vlr_Descontos': parse_br_number(row[5].value),
                'Ticket_Medio': parse_br_number(row[6].value),
                'Vlr_Venda': vlr_venda,
                'Part_Venda': parse_br_number(row[8].value),
                'Markdown_Pct': parse_br_number(row[9].value),
                'Markdown_Ult_Entrada': parse_br_number(row[10].value),
                'Markup_Pct': parse_br_number(row[11].value),
                'Markup_Ult_Entrada': parse_br_number(row[12].value),
                'Vlr_Lucro': vlr_lucro,
                'Part_Lucro': parse_br_number(row[14].value),
                'Custo_Medio_Liq': parse_br_number(row[15].value),
                'Custo_Ult_Entrada_Liq': parse_br_number(row[16].value),
                'Margem_Pct': round(margem, 2),
            })
    
    wb.close()
    return vendas_diarias


# ============================================================
# 3. PROCESSAR CURVA A
# ============================================================
def processar_curva_a(filepath, mes, ano):
    """Processa arquivo da Curva A e retorna dados dos produtos vitais"""
    wb = openpyxl.load_workbook(filepath)
    ws = wb['Main sheet']
    
    produtos_curva_a = []
    
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=False):
        col_a = row[0].value
        if col_a is None or col_a == 'Total':
            continue
        
        prod_parts = str(col_a).split('||')
        prod_name = prod_parts[0].strip()
        prod_code = prod_parts[1].strip() if len(prod_parts) > 1 else ''
        
        vlr_venda = parse_br_number(row[6].value)
        vlr_lucro = parse_br_number(row[12].value)
        margem = (vlr_lucro / vlr_venda * 100) if vlr_venda > 0 else 0
        
        produtos_curva_a.append({
            'Mes': mes,
            'Ano': ano,
            'Periodo': f"{mes:02d}/{ano}",
            'Produto': prod_name,
            'Codigo': prod_code,
            'Qtde_Venda': parse_br_number(row[1].value),
            'Qtde_Documentos': parse_br_number(row[2].value),
            'Vlr_Venda': vlr_venda,
            'Markdown_Pct': parse_br_number(row[8].value),
            'Markdown_Ult_Entrada': parse_br_number(row[9].value),
            'Markup_Pct': parse_br_number(row[10].value),
            'Markup_Ult_Entrada': parse_br_number(row[11].value),
            'Vlr_Lucro': vlr_lucro,
            'Custo_Medio_Liq': parse_br_number(row[14].value),
            'Custo_Ult_Entrada_Liq': parse_br_number(row[15].value),
            'Margem_Pct': round(margem, 2),
            'Erosao_Margem': round(parse_br_number(row[8].value) - parse_br_number(row[9].value), 2),
        })
    
    wb.close()
    return produtos_curva_a


# ============================================================
# 4. PROCESSAR HISTÓRICO 2025
# ============================================================
def processar_historico_2025(filepath):
    """Processa o arquivo de histórico anual 2025"""
    wb = openpyxl.load_workbook(filepath)
    ws = wb['Main sheet']
    
    historico = []
    current_month = None
    
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=False):
        col_a = row[0].value
        col_b = row[1].value
        
        if col_a is not None and col_a != 'Total':
            month_lower = str(col_a).strip().lower()
            if month_lower in MESES_PT:
                current_month = month_lower
        elif col_b is not None and 'Total' not in str(col_b) and current_month:
            prod_parts = str(col_b).split('||')
            prod_name = prod_parts[0].strip()
            
            vlr_venda = parse_br_number(row[7].value)
            vlr_lucro = parse_br_number(row[13].value)
            
            # Filtrar anomalias óbvias (lucro < -1000 com receita < 500)
            if vlr_lucro < -1000 and vlr_venda < 500:
                continue
            
            margem = (vlr_lucro / vlr_venda * 100) if vlr_venda > 0 else 0
            
            historico.append({
                'Mes': MESES_PT[current_month],
                'Ano': 2025,
                'Periodo': f"{MESES_PT[current_month]:02d}/2025",
                'Nome_Mes': current_month.capitalize(),
                'Produto': prod_name,
                'Qtde_Venda': parse_br_number(row[2].value),
                'Qtde_Documentos': parse_br_number(row[3].value),
                'Vlr_Venda': vlr_venda,
                'Vlr_Lucro': vlr_lucro,
                'Margem_Pct': round(margem, 2),
                'Markdown_Pct': parse_br_number(row[9].value),
            })
    
    wb.close()
    return historico


# ============================================================
# 5. CALCULAR MÉTRICAS DE PRODUTO
# ============================================================
def calcular_metricas_produto(vendas_diarias, curva_a_nomes, dias_operacao):
    """Calcula giro, classificação na matriz e alertas de ruptura"""
    
    # Agregar por produto
    produtos = defaultdict(lambda: {
        'dias_vendidos': 0, 'receita': 0, 'lucro': 0, 'qtde': 0,
        'cupons': 0, 'datas': [], 'margens': []
    })
    
    for v in vendas_diarias:
        p = produtos[v['Produto']]
        p['dias_vendidos'] += 1
        p['receita'] += v['Vlr_Venda']
        p['lucro'] += v['Vlr_Lucro']
        p['qtde'] += v['Qtde_Venda']
        p['cupons'] += v['Qtde_Documentos']
        p['datas'].append(v['Data'])
        if v['Vlr_Venda'] > 0:
            p['margens'].append(v['Vlr_Lucro'] / v['Vlr_Venda'] * 100)
    
    resultado = []
    for nome, data in produtos.items():
        giro = data['dias_vendidos'] / dias_operacao if dias_operacao > 0 else 0
        margem_media = sum(data['margens']) / len(data['margens']) if data['margens'] else 0
        
        # Classificação na Matriz 2x2
        if giro > LIMIAR_GIRO_ALTO and margem_media > LIMIAR_MARGEM_ALTA:
            classificacao = "⭐ Estrela"
        elif giro > LIMIAR_GIRO_ALTO and margem_media <= LIMIAR_MARGEM_ALTA:
            classificacao = "💰 Gerador de Caixa"
        elif giro <= LIMIAR_GIRO_ALTO and margem_media > LIMIAR_MARGEM_ALTA:
            classificacao = "🔍 Oportunidade"
        else:
            classificacao = "⚠️ Peso Morto"
        
        # Curva
        curva = "A" if nome in curva_a_nomes else "B/C"
        
        # Alerta de ruptura: produto de giro diário sem venda nos últimos 2 dias úteis
        alerta_ruptura = False
        if giro >= LIMIAR_RUPTURA_GIRO:
            datas_sorted = sorted(data['datas'])
            if datas_sorted:
                ultima_venda = datas_sorted[-1]
                # Verificar se faz mais de 2 dias úteis desde a última venda
                # (essa lógica será refinada com dados reais)
                alerta_ruptura = data['dias_vendidos'] >= (dias_operacao * LIMIAR_RUPTURA_GIRO) and len(datas_sorted) > 0
        
        resultado.append({
            'Produto': nome,
            'Curva': curva,
            'Dias_Vendidos': data['dias_vendidos'],
            'Dias_Operacao': dias_operacao,
            'Giro': round(giro, 3),
            'Receita_Total': round(data['receita'], 2),
            'Lucro_Total': round(data['lucro'], 2),
            'Margem_Media': round(margem_media, 2),
            'Qtde_Total': round(data['qtde'], 3),
            'Cupons_Total': round(data['cupons'], 0),
            'Classificacao': classificacao,
            'Receita_Media_Dia': round(data['receita'] / data['dias_vendidos'], 2) if data['dias_vendidos'] > 0 else 0,
            'Giro_Diario': "Sim" if giro >= LIMIAR_RUPTURA_GIRO else "Não",
        })
    
    return resultado


# ============================================================
# 6. GERAR COMPARATIVO YOY
# ============================================================
def gerar_comparativo_yoy(historico_2025, categorias_2026, mes_2026, ano_2026):
    """Gera comparativo mês a mês entre 2025 e 2026"""
    
    # Agregar histórico 2025 por mês
    hist_mensal = defaultdict(lambda: {'receita': 0, 'lucro': 0, 'cupons': 0, 'produtos': 0})
    for h in historico_2025:
        key = h['Mes']
        hist_mensal[key]['receita'] += h['Vlr_Venda']
        hist_mensal[key]['lucro'] += h['Vlr_Lucro']
        hist_mensal[key]['cupons'] += h['Qtde_Documentos']
        hist_mensal[key]['produtos'] += 1
    
    comparativo = []
    meses_nomes = ['', 'Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
                   'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']
    
    for mes_num in range(1, 13):
        h = hist_mensal.get(mes_num, {'receita': 0, 'lucro': 0, 'cupons': 0, 'produtos': 0})
        margem_25 = (h['lucro'] / h['receita'] * 100) if h['receita'] > 0 else 0
        
        row = {
            'Mes': meses_nomes[mes_num],
            'Mes_Num': mes_num,
            'Receita_2025': round(h['receita'], 2),
            'Lucro_2025': round(h['lucro'], 2),
            'Margem_2025': round(margem_25, 2),
            'Cupons_2025': round(h['cupons'], 0),
            'SKUs_2025': h['produtos'],
            'Receita_2026': 0,
            'Lucro_2026': 0,
            'Margem_2026': 0,
            'Cupons_2026': 0,
            'Var_Receita_Pct': 0,
            'Var_Lucro_Pct': 0,
        }
        comparativo.append(row)
    
    # Preencher dados de 2026 disponíveis
    if categorias_2026:
        total_receita_26 = sum(c['Vlr_Venda'] for c in categorias_2026)
        total_lucro_26 = sum(c['Vlr_Lucro'] for c in categorias_2026)
        total_cupons_26 = sum(c['Qtde_Documentos'] for c in categorias_2026)
        margem_26 = (total_lucro_26 / total_receita_26 * 100) if total_receita_26 > 0 else 0
        
        for row in comparativo:
            if row['Mes_Num'] == mes_2026:
                row['Receita_2026'] = round(total_receita_26, 2)
                row['Lucro_2026'] = round(total_lucro_26, 2)
                row['Margem_2026'] = round(margem_26, 2)
                row['Cupons_2026'] = round(total_cupons_26, 0)
                if row['Receita_2025'] > 0:
                    row['Var_Receita_Pct'] = round((total_receita_26 - row['Receita_2025']) / row['Receita_2025'] * 100, 2)
                if row['Lucro_2025'] > 0:
                    row['Var_Lucro_Pct'] = round((total_lucro_26 - row['Lucro_2025']) / row['Lucro_2025'] * 100, 2)
    
    return comparativo


# ============================================================
# 7. GERAR CALENDÁRIO
# ============================================================
def gerar_calendario(ano):
    """Gera dimensão calendário para o ano"""
    dias_semana = {0: 'Segunda', 1: 'Terça', 2: 'Quarta', 3: 'Quinta',
                   4: 'Sexta', 5: 'Sábado', 6: 'Domingo'}
    meses_nomes = ['', 'Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
                   'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']
    
    calendario = []
    start = datetime(ano, 1, 1)
    end = datetime(ano, 12, 31)
    current = start
    
    while current <= end:
        weekday = current.weekday()
        is_feriado = current.strftime('%Y-%m-%d') in FERIADOS_SET
        is_domingo = weekday == 6
        is_util = weekday < 6 and not is_feriado  # Seg-Sáb e não feriado
        
        # Semana do mês
        day = current.day
        semana_mes = (day - 1) // 7 + 1
        
        calendario.append({
            'Data': current.strftime('%Y-%m-%d'),
            'Dia': current.day,
            'Dia_Semana': dias_semana[weekday],
            'Dia_Semana_Num': weekday + 1,
            'Semana_Mes': semana_mes,
            'Mes': current.month,
            'Nome_Mes': meses_nomes[current.month],
            'Ano': current.year,
            'Trimestre': f"Q{(current.month - 1) // 3 + 1}",
            'E_Util': "Sim" if is_util else "Não",
            'E_Domingo': "Sim" if is_domingo else "Não",
            'E_Feriado': "Sim" if is_feriado else "Não",
        })
        current += timedelta(days=1)
    
    return calendario


# ============================================================
# 8. ESCREVER EXCEL FINAL
# ============================================================
def escrever_excel(output_path, categorias, vendas_diarias, dim_produtos, 
                   dim_calendario, comparativo_yoy, curva_a, alertas_erosao):
    """Gera o arquivo Base_PowerBI.xlsx com todas as tabelas"""
    
    wb = openpyxl.Workbook()
    
    # --- ABA 1: fato_vendas_mensais ---
    ws1 = wb.active
    ws1.title = "fato_vendas_mensais"
    headers = ['Periodo', 'Mes', 'Ano', 'Categoria', 'Qtde_Venda', 'Qtde_Documentos',
               'Ticket_Medio', 'Vlr_Venda', 'Markdown_Pct', 'Markdown_Ult_Entrada',
               'Markup_Pct', 'Markup_Ult_Entrada', 'Vlr_Lucro', 'Custo_Medio_Liq',
               'Custo_Ult_Entrada_Liq']
    ws1.append(headers)
    for cat in categorias:
        if cat['Categoria'] != 'Total':
            ws1.append([cat.get(h, '') for h in headers])
    style_header(ws1)
    style_data_rows(ws1)
    auto_width(ws1)
    
    # --- ABA 2: fato_vendas_diarias ---
    ws2 = wb.create_sheet("fato_vendas_diarias")
    headers2 = ['Data', 'Periodo', 'Produto', 'Codigo', 'Qtde_Venda', 'Qtde_Documentos',
                'Vlr_Venda', 'Vlr_Lucro', 'Margem_Pct', 'Markdown_Pct', 
                'Markdown_Ult_Entrada', 'Markup_Pct', 'Markup_Ult_Entrada',
                'Custo_Medio_Liq', 'Custo_Ult_Entrada_Liq']
    ws2.append(headers2)
    for v in vendas_diarias:
        ws2.append([v.get(h, '') for h in headers2])
    style_header(ws2)
    style_data_rows(ws2)
    auto_width(ws2)
    
    # --- ABA 3: dim_produtos ---
    ws3 = wb.create_sheet("dim_produtos")
    headers3 = ['Produto', 'Curva', 'Classificacao', 'Dias_Vendidos', 'Dias_Operacao',
                'Giro', 'Receita_Total', 'Lucro_Total', 'Margem_Media', 'Qtde_Total',
                'Cupons_Total', 'Receita_Media_Dia', 'Giro_Diario']
    ws3.append(headers3)
    for p in sorted(dim_produtos, key=lambda x: -x['Receita_Total']):
        ws3.append([p.get(h, '') for h in headers3])
    
    style_header(ws3)
    style_data_rows(ws3)
    
    # Colorir classificações
    for row_idx in range(2, ws3.max_row + 1):
        cell = ws3.cell(row=row_idx, column=3)  # Classificacao
        val = str(cell.value)
        if 'Estrela' in val:
            cell.fill = PatternFill('solid', fgColor='D5F5E3')
        elif 'Gerador' in val:
            cell.fill = PatternFill('solid', fgColor='FFF9C4')
        elif 'Oportunidade' in val:
            cell.fill = PatternFill('solid', fgColor='DCEEFB')
        elif 'Peso Morto' in val:
            cell.fill = PatternFill('solid', fgColor='FADBD8')
    
    auto_width(ws3)
    
    # --- ABA 4: dim_calendario ---
    ws4 = wb.create_sheet("dim_calendario")
    headers4 = ['Data', 'Dia', 'Dia_Semana', 'Dia_Semana_Num', 'Semana_Mes',
                'Mes', 'Nome_Mes', 'Ano', 'Trimestre', 'E_Util', 'E_Domingo', 'E_Feriado']
    ws4.append(headers4)
    for d in dim_calendario:
        ws4.append([d.get(h, '') for h in headers4])
    style_header(ws4)
    style_data_rows(ws4)
    auto_width(ws4)
    
    # --- ABA 5: comparativo_yoy ---
    ws5 = wb.create_sheet("comparativo_yoy")
    headers5 = ['Mes', 'Mes_Num', 'Receita_2025', 'Lucro_2025', 'Margem_2025',
                'Cupons_2025', 'SKUs_2025', 'Receita_2026', 'Lucro_2026', 
                'Margem_2026', 'Cupons_2026', 'Var_Receita_Pct', 'Var_Lucro_Pct']
    ws5.append(headers5)
    for c in comparativo_yoy:
        ws5.append([c.get(h, '') for h in headers5])
    style_header(ws5)
    style_data_rows(ws5)
    auto_width(ws5)
    
    # --- ABA 6: alertas_erosao_margem ---
    ws6 = wb.create_sheet("alertas_erosao_margem")
    headers6 = ['Produto', 'Periodo', 'Curva', 'Vlr_Venda', 'Vlr_Lucro', 'Margem_Pct',
                'Markdown_Pct', 'Markdown_Ult_Entrada', 'Erosao_Margem', 'Alerta']
    ws6.append(headers6)
    for a in sorted(alertas_erosao, key=lambda x: x['Erosao_Margem']):
        alerta = "🔴 CUSTO SUBIU" if a['Erosao_Margem'] > LIMIAR_EROSAO else (
                 "🟢 CUSTO CAIU" if a['Erosao_Margem'] < -LIMIAR_EROSAO else "⚪ Estável")
        ws6.append([
            a['Produto'], a['Periodo'], 'A', a['Vlr_Venda'], a['Vlr_Lucro'],
            a['Margem_Pct'], a['Markdown_Pct'], a['Markdown_Ult_Entrada'],
            a['Erosao_Margem'], alerta
        ])
    style_header(ws6)
    style_data_rows(ws6)
    
    # Colorir alertas
    for row_idx in range(2, ws6.max_row + 1):
        cell = ws6.cell(row=row_idx, column=10)  # Alerta
        val = str(cell.value)
        if 'SUBIU' in val:
            cell.fill = PatternFill('solid', fgColor='FADBD8')
        elif 'CAIU' in val:
            cell.fill = PatternFill('solid', fgColor='D5F5E3')
    
    auto_width(ws6)
    
    # --- ABA RESUMO ---
    ws_resumo = wb.create_sheet("resumo_executivo")
    wb.move_sheet(ws_resumo, offset=-6)  # Mover para primeira posição
    
    ws_resumo.column_dimensions['A'].width = 35
    ws_resumo.column_dimensions['B'].width = 25
    ws_resumo.column_dimensions['C'].width = 25
    ws_resumo.column_dimensions['D'].width = 20
    
    # Título
    ws_resumo['A1'] = 'MERCADO duBAIRRO — Resumo Executivo'
    ws_resumo['A1'].font = Font(name='Arial', bold=True, size=16, color='2D2D2D')
    ws_resumo['A2'] = 'Painel dos Sócios'
    ws_resumo['A2'].font = Font(name='Arial', size=12, color='666666')
    
    # Calcular KPIs do mês atual
    total_receita = sum(c['Vlr_Venda'] for c in categorias if c['Categoria'] != 'Total')
    total_lucro = sum(c['Vlr_Lucro'] for c in categorias if c['Categoria'] != 'Total')
    lucro_liquido = total_lucro - CUSTO_FIXO
    margem_bruta = (total_lucro / total_receita * 100) if total_receita > 0 else 0
    margem_real = (lucro_liquido / total_receita * 100) if total_receita > 0 else 0
    ponto_equilibrio = CUSTO_FIXO / (margem_bruta / 100) if margem_bruta > 0 else 0
    total_cupons = sum(c['Qtde_Documentos'] for c in categorias if c['Categoria'] != 'Total')
    ticket_medio = total_receita / total_cupons if total_cupons > 0 else 0
    
    # Cards KPI
    kpis = [
        ('', '', '', ''),
        ('KPI', 'Valor', 'Meta / Referência', 'Status'),
        ('Faturamento', f'R$ {total_receita:,.2f}', '', ''),
        ('Lucro Bruto', f'R$ {total_lucro:,.2f}', '', ''),
        ('Lucro Líquido (- Custo Fixo)', f'R$ {lucro_liquido:,.2f}', f'Meta: R$ {total_receita * META_LIQUIDA:,.2f}', 
         '✅ Acima' if lucro_liquido > total_receita * META_LIQUIDA else '⚠️ Abaixo'),
        ('Margem Bruta', f'{margem_bruta:.1f}%', '', ''),
        ('Margem Real', f'{margem_real:.1f}%', f'Meta: {META_LIQUIDA*100:.0f}%',
         '✅ Saudável' if margem_real > META_LIQUIDA * 100 else '🔴 Atenção'),
        ('Ponto de Equilíbrio', f'R$ {ponto_equilibrio:,.2f}', f'Folga: {((total_receita/ponto_equilibrio)-1)*100:.0f}%' if ponto_equilibrio > 0 else '', ''),
        ('Nº de Cupons', f'{total_cupons:,.0f}', '', ''),
        ('Ticket Médio', f'R$ {ticket_medio:.2f}', '', ''),
        ('Custo Fixo Mensal', f'R$ {CUSTO_FIXO:,.2f}', '', ''),
        ('Produtos Ativos (SKUs)', f'{len(set(v["Produto"] for v in vendas_diarias)):,}', '', ''),
        ('Produtos Curva A', f'{len(curva_a)}', '', ''),
    ]
    
    for row_idx, (k, v, ref, status) in enumerate(kpis, start=4):
        ws_resumo.cell(row=row_idx, column=1, value=k)
        ws_resumo.cell(row=row_idx, column=2, value=v)
        ws_resumo.cell(row=row_idx, column=3, value=ref)
        ws_resumo.cell(row=row_idx, column=4, value=status)
    
    # Header da tabela KPI
    for col in range(1, 5):
        cell = ws_resumo.cell(row=5, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
    
    for row_idx in range(6, 17):
        for col in range(1, 5):
            cell = ws_resumo.cell(row=row_idx, column=col)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            if row_idx % 2 == 0:
                cell.fill = LIGHT_GRAY_FILL
    
    wb.save(output_path)
    return output_path


# ============================================================
# MAIN
# ============================================================
def main():
    print("=" * 60)
    print("MERCADO duBAIRRO — Processamento de Dados")
    print("=" * 60)
    
    # Detectar arquivos
    base_path = "/mnt/user-data/uploads"
    
    # Encontrar arquivo de categorias
    cat_files = glob.glob(os.path.join(base_path, "categoria_analisedevendas_*.xlsx"))
    prod_files = glob.glob(os.path.join(base_path, "produtopordia_analisedevendas_*.xlsx"))
    curva_files = glob.glob(os.path.join(base_path, "curvaA_analisedevendas_*.xlsx"))
    hist_file = os.path.join(base_path, "mesamesproduto2025_analisedevendas.xlsx")
    
    if not cat_files:
        print("ERRO: Arquivo de categorias não encontrado!")
        sys.exit(1)
    if not prod_files:
        print("ERRO: Arquivo de produto por dia não encontrado!")
        sys.exit(1)
    if not curva_files:
        print("ERRO: Arquivo de curva A não encontrado!")
        sys.exit(1)
    
    # Processar cada mês disponível
    all_categorias = []
    all_vendas = []
    all_curva_a = []
    
    for f in cat_files:
        mes, ano = detect_month_year(f)
        if mes and ano:
            print(f"\n📁 Processando categorias: {os.path.basename(f)} ({mes}/{ano})")
            cats, total = processar_categorias(f, mes, ano)
            all_categorias.extend(cats)
            print(f"   → {len(cats)} categorias processadas")
    
    # Usar o primeiro mês encontrado para referência
    mes_ref = all_categorias[0]['Mes'] if all_categorias else 1
    ano_ref = all_categorias[0]['Ano'] if all_categorias else 2026
    
    for f in prod_files:
        mes, ano = detect_month_year(f)
        if mes and ano:
            print(f"\n📁 Processando produto/dia: {os.path.basename(f)} ({mes}/{ano})")
            vendas = processar_produto_dia(f, mes, ano)
            all_vendas.extend(vendas)
            print(f"   → {len(vendas)} registros de venda diária")
            
            # Contar dias de operação
            datas_unicas = set(v['Data'] for v in vendas)
            dias_op = len(datas_unicas)
            print(f"   → {dias_op} dias de operação detectados")
    
    for f in curva_files:
        mes, ano = detect_month_year(f)
        if mes and ano:
            print(f"\n📁 Processando Curva A: {os.path.basename(f)} ({mes}/{ano})")
            curva = processar_curva_a(f, mes, ano)
            all_curva_a.extend(curva)
            print(f"   → {len(curva)} produtos Curva A")
    
    # Processar histórico
    historico = []
    if os.path.exists(hist_file):
        print(f"\n📁 Processando histórico 2025...")
        historico = processar_historico_2025(hist_file)
        print(f"   → {len(historico)} registros históricos")
    
    # Calcular métricas de produto
    print(f"\n🔧 Calculando métricas de produto...")
    curva_a_nomes = set(c['Produto'] for c in all_curva_a)
    datas_unicas = set(v['Data'] for v in all_vendas)
    dias_op = len(datas_unicas)
    dim_produtos = calcular_metricas_produto(all_vendas, curva_a_nomes, dias_op)
    
    # Estatísticas da matriz
    estrelas = sum(1 for p in dim_produtos if 'Estrela' in p['Classificacao'])
    geradores = sum(1 for p in dim_produtos if 'Gerador' in p['Classificacao'])
    oportunidades = sum(1 for p in dim_produtos if 'Oportunidade' in p['Classificacao'])
    peso_morto = sum(1 for p in dim_produtos if 'Peso Morto' in p['Classificacao'])
    print(f"   ⭐ Estrelas: {estrelas}")
    print(f"   💰 Geradores de Caixa: {geradores}")
    print(f"   🔍 Oportunidades: {oportunidades}")
    print(f"   ⚠️  Peso Morto: {peso_morto}")
    
    # Gerar comparativo YoY
    print(f"\n🔧 Gerando comparativo YoY...")
    comparativo = gerar_comparativo_yoy(historico, all_categorias, mes_ref, ano_ref)
    
    # Gerar calendário
    print(f"\n🔧 Gerando calendário 2026...")
    calendario = gerar_calendario(2026)
    # Adicionar 2025 também
    calendario.extend(gerar_calendario(2025))
    calendario.sort(key=lambda x: x['Data'])
    
    # Gerar alertas de erosão
    alertas_erosao = [c for c in all_curva_a if abs(c['Erosao_Margem']) > 0]
    
    # Escrever Excel
    output_path = "/mnt/user-data/outputs/Base_PowerBI.xlsx"
    print(f"\n💾 Gerando Base_PowerBI.xlsx...")
    escrever_excel(output_path, all_categorias, all_vendas, dim_produtos,
                   calendario, comparativo, all_curva_a, alertas_erosao)
    
    print(f"\n{'=' * 60}")
    print(f"✅ CONCLUÍDO!")
    print(f"   Arquivo: {output_path}")
    print(f"   Abas geradas:")
    print(f"     1. resumo_executivo     → KPIs principais")
    print(f"     2. fato_vendas_mensais  → {len(all_categorias)} registros")
    print(f"     3. fato_vendas_diarias  → {len(all_vendas)} registros")
    print(f"     4. dim_produtos         → {len(dim_produtos)} produtos classificados")
    print(f"     5. dim_calendario       → {len(calendario)} dias")
    print(f"     6. comparativo_yoy      → 12 meses comparados")
    print(f"     7. alertas_erosao_margem→ {len(alertas_erosao)} produtos monitorados")
    print(f"{'=' * 60}")


if __name__ == "__main__":
    main()
