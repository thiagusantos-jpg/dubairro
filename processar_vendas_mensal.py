import pandas as pd
from openpyxl import load_workbook
import os

def limpar_financeiro(v):
    """Transforma texto do sistema em número real sem multiplicar por 100."""
    if pd.isna(v): return 0.0
    if isinstance(v, (int, float)): return float(v)
    
    s = str(v).replace('R$', '').strip()
    
    # Se tiver os dois (1.234,56), tira o ponto e troca a vírgula
    if ',' in s and '.' in s:
        s = s.replace('.', '').replace(',', '.')
    # Se tiver só vírgula (32,37), vira ponto (32.37)
    elif ',' in s:
        s = s.replace(',', '.')
    # Se tiver só ponto (32.37), o Python já entende, não fazemos nada
    
    try:
        return float(s)
    except:
        return 0.0

def processar():
    print("🧹 LIMPANDO E CORRIGINDO VALORES DE JANEIRO...")
    
    if not os.path.exists('JANEIRO2026.xlsx'):
        print("❌ Arquivo JANEIRO2026.xlsx não encontrado!")
        return

    # 1. Lê os dados
    df = pd.read_excel('JANEIRO2026.xlsx', skiprows=1)

    # 2. Extrai Nome e Código Corretamente
    df['Nome_Original'] = df.iloc[:, 0].astype(str)
    df['Produto'] = df['Nome_Original'].str.split('||', regex=False).str[0].str.strip()
    df['Codigo']  = df['Nome_Original'].str.split('||', regex=False).str[1].str.strip()

    # 3. Converte os números com a lógica correta
    df['Qtde'] = df['Qtde. Venda'].apply(limpar_financeiro) / 1000
    df['Venda_Total'] = df['Vlr. Venda'].apply(limpar_financeiro)
    df['Custo_Total'] = df['Custo Médio Líq.'].apply(limpar_financeiro)

    # Filtra apenas o que é venda real
    df = df[(df['Venda_Total'] > 0) & (df['Produto'] != 'nan') & (df['Produto'] != 'Total')]

    # 4. Salva na Base_PowerBI.xlsx
    wb = load_workbook('Base_PowerBI.xlsx')
    ws = wb['DadosVendas']
    
    # Limpa dados antigos antes de gravar os novos
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row)

    for _, r in df.iterrows():
        ws.append([
            '2026-01-01', r['Produto'], r['Codigo'], 'Geral',
            r['Qtde'], (r['Venda_Total']/r['Qtde']) if r['Qtde'] > 0 else 0, 
            (r['Custo_Total']/r['Qtde']) if r['Qtde'] > 0 else 0,
            r['Venda_Total'], r['Custo_Total'], 
            (r['Venda_Total'] - r['Custo_Total']) / r['Venda_Total'] if r['Venda_Total'] > 0 else 0
        ])
    
    wb.save('Base_PowerBI.xlsx')
    
    print(f"\n✅ SUCESSO!")
    print(f"💰 Faturamento Real: R$ {df['Venda_Total'].sum():,.2f}")
    print(f"📦 Itens no Dashboard: {len(df)}")

if __name__ == "__main__":
    processar()